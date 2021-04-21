from openpyxl import load_workbook
import sys

book = load_workbook("blue.xlsx")
sheet = book.active
ultima_fecha = sheet["A"+str(sheet.max_row)].value

book_aux = load_workbook("blue_2021.xlsx")
sheet_aux = book_aux.active
ultima_posicion_aux = sheet_aux.max_row


for posicion, row in enumerate(sheet_aux):
    if row[0].value == ultima_fecha:
        if posicion == ultima_posicion_aux-1:
            print("HOJA YA ACTUALIZADA")
            sys.exit()
        break
else:
    sys.exit()

tabla = []
for row in sheet_aux["a"+str(posicion+2):"c"+str(ultima_posicion_aux)]:
    tabla.append((row[0].value,row[1].value,row[2].value))

print("Se agregaron las siguientes tablas")
for row in tabla:
    print(row)
    sheet.append(row)

book.save('blue.xlsx')
print("GUARDADO CORRECTAMENTE")